# main.py
"""
EY Project — Work Order Extraction Pipeline
PDF → Azure Document Intelligence → Rules + LLM → Formatted Excel
"""

import json
import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.azure_di import analyze_pdf
from src.rule_extractor import extract_workorder
from src.llm_fallback import should_use_llm, enhance_with_llm
from src.config import OUTPUT_DIR, USE_LLM_FALLBACK, DI_MODEL


# ─────────────────────────────────────────────────────────────
# Excel formatting helpers
# ─────────────────────────────────────────────────────────────

_DARK_BLUE  = "1F3864"
_MID_BLUE   = "2E5FAA"
_LIGHT_GRAY = "F5F5F5"
_WHITE      = "FFFFFF"
_GREEN_BG   = "E2EFDA"
_GREEN_FG   = "375623"


def _hdr_font():
    return Font(name="Arial", bold=True, color=_WHITE, size=10)

def _body_font(bold=False, color="000000", italic=False):
    return Font(name="Arial", bold=bold, color=color, size=10, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(top=s, bottom=s, left=s, right=s)

def _apply_header_row(ws, row: int, headers: list):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _hdr_font()
        c.fill      = _fill(_DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin_border()
    ws.row_dimensions[row].height = 20

def _apply_data_row(ws, row: int, values: list, alt: bool = False):
    bg = _LIGHT_GRAY if alt else _WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=str(v) if v is not None else "")
        c.font      = _body_font()
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = _thin_border()

def _sheet_banner(ws, title: str, num_cols: int):
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Arial", bold=True, size=13, color=_WHITE)
    c.fill      = _fill(_MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────

def _write_header_sheet(wb: Workbook, header: dict):
    ws = wb.create_sheet("Header")
    ws.sheet_view.showGridLines = False
    _sheet_banner(ws, "Work Order — Header Information", 2)
    _apply_header_row(ws, 2, ["Field", "Value"])
    for i, (field, value) in enumerate(header.items()):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        _apply_data_row(ws, r, [field, value], alt=i % 2 == 0)
        ws.cell(r, 1).font = _body_font(bold=True)
    _set_col_widths(ws, [32, 55])
    ws.freeze_panes = "A3"


def _write_services_sheet(wb: Workbook, services: list):
    ws = wb.create_sheet("Services")
    ws.sheet_view.showGridLines = False
    cols = ["Sr No", "SrvLnNo", "SrvNo", "Brief Description", "Long Text", "Rate", "Unit"]
    _sheet_banner(ws, "Service Line Items", len(cols))
    _apply_header_row(ws, 2, cols)
    for i, svc in enumerate(services):
        r = 3 + i
        ws.row_dimensions[r].height = 60
        _apply_data_row(ws, r, [svc.get(c, "") for c in cols], alt=i % 2 == 0)
    _set_col_widths(ws, [8, 10, 14, 30, 65, 10, 22])
    ws.freeze_panes = "A3"


def _write_pricing_sheet(wb: Workbook, pricing: dict):
    ws = wb.create_sheet("Pricing")
    ws.sheet_view.showGridLines = False
    _sheet_banner(ws, "Pricing & Rate Information", 2)
    _apply_header_row(ws, 2, ["Field", "Value"])
    for i, (field, value) in enumerate(pricing.items()):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        _apply_data_row(ws, r, [field, value], alt=i % 2 == 0)
        ws.cell(r, 1).font = _body_font(bold=True)
        if any(x in field.lower() for x in ["price", "rate", "ceiling", "value"]):
            ws.cell(r, 2).fill = _fill(_GREEN_BG)
            ws.cell(r, 2).font = _body_font(bold=True, color=_GREEN_FG)
    _set_col_widths(ws, [32, 40])
    ws.freeze_panes = "A3"


def _write_text_blocks_sheet(wb: Workbook, text_blocks: dict):
    ws = wb.create_sheet("Text Blocks")
    ws.sheet_view.showGridLines = False
    _sheet_banner(ws, "Extracted Text Blocks", 2)
    _apply_header_row(ws, 2, ["Section", "Content"])
    for i, (section, content) in enumerate(text_blocks.items()):
        r = 3 + i
        ws.row_dimensions[r].height = 80
        _apply_data_row(ws, r, [section, content], alt=i % 2 == 0)
        ws.cell(r, 1).font = _body_font(bold=True)
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    _set_col_widths(ws, [28, 120])
    ws.freeze_panes = "A3"


def _write_change_orders_sheet(wb: Workbook, change_orders: list):
    ws = wb.create_sheet("Change Orders")
    ws.sheet_view.showGridLines = False
    cols = ["C/O Date", "Amendment Type", "Description", "New Validity", "Ceiling Change"]
    _sheet_banner(ws, "Contract Change Orders (Amendments)", len(cols))
    _apply_header_row(ws, 2, cols)
    if not change_orders:
        c = ws.cell(row=3, column=1, value="No change orders found in document")
        c.font = _body_font(italic=True, color="888888")
    else:
        for i, co in enumerate(change_orders):
            r = 3 + i
            ws.row_dimensions[r].height = 45
            _apply_data_row(ws, r, [co.get(c, "") for c in cols], alt=i % 2 == 0)
    _set_col_widths(ws, [14, 22, 80, 16, 20])
    ws.freeze_panes = "A3"


def _write_metadata_sheet(wb: Workbook, metadata: dict, pdf_path: str):
    ws = wb.create_sheet("Metadata")
    ws.sheet_view.showGridLines = False
    _sheet_banner(ws, "Extraction Metadata", 2)
    _apply_header_row(ws, 2, ["Property", "Value"])
    rows = [
        ("Source PDF",      os.path.basename(pdf_path)),
        ("Extraction Time", metadata.get("extracted_at", "")),
        ("DI Model",        metadata.get("model", "")),
        ("Pages Analyzed",  str(metadata.get("pages", ""))),
        ("Paragraphs",      str(metadata.get("paragraphs", ""))),
        ("KV Pairs Found",  str(metadata.get("kv_pairs", ""))),
        ("Tables Found",    str(metadata.get("tables", ""))),
    ]
    for i, (k, v) in enumerate(rows):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        _apply_data_row(ws, r, [k, v], alt=i % 2 == 0)
        ws.cell(r, 1).font = _body_font(bold=True)
    _set_col_widths(ws, [28, 50])


# ─────────────────────────────────────────────────────────────
# Save to Excel
# ─────────────────────────────────────────────────────────────

def save_to_excel(data: dict, pdf_path: str) -> str:
    base      = Path(pdf_path).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = os.path.join(OUTPUT_DIR, f"{base}_extracted_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _write_header_sheet(wb,        data.get("header", {}))
    _write_services_sheet(wb,      data.get("services", []))
    _write_pricing_sheet(wb,       data.get("pricing", {}))
    _write_text_blocks_sheet(wb,   data.get("text_blocks", {}))
    _write_change_orders_sheet(wb, data.get("change_orders", []))
    _write_metadata_sheet(wb,      data.get("metadata", {}), pdf_path)

    wb.save(out_path)
    print(f"✓ Excel saved: {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
# Pipeline
# ─────────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> str:
    print(f"\n{'═' * 72}")
    print(f"  EY — Work Order Extraction Pipeline")
    print(f"  File  : {pdf_path}")
    print(f"  Model : {DI_MODEL}  |  LLM: {'ON' if USE_LLM_FALLBACK else 'OFF'}")
    print(f"  Time  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'═' * 72}\n")

    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    print("Step 1/3  Azure Document Intelligence...")
    result = analyze_pdf(pdf_path)

    print("Step 2/3  Rule-based extraction (full PDF via pypdf)...")
    data = extract_workorder(result, pdf_path)

    if USE_LLM_FALLBACK:
        print("Step 3/3  LLM gap-fill check...")
        if should_use_llm(data):
            from src.pdf_extractor import extract_text_from_pdf
            raw_text = extract_text_from_pdf(pdf_path)
            data = enhance_with_llm(raw_text, data)
        else:
            print("          No gaps — skipping LLM")
    else:
        print("Step 3/3  LLM disabled")

    print(f"\n  ── Summary ──")
    print(f"  Header : {len(data.get('header', {}))} fields")
    print(f"  Services: {len(data.get('services', []))} items")
    print(f"  C/O    : {len(data.get('change_orders', []))} amendments")
    print(f"  Pricing: {len(data.get('pricing', {}))} fields")

    excel_path = save_to_excel(data, pdf_path)
    print(f"\n{'═' * 72}")
    print(f"  ✓ Done — {excel_path}")
    print(f"{'═' * 72}\n")
    return excel_path


def process_folder(folder_path: str) -> list:
    pdfs = list(Path(folder_path).glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {folder_path}")
        return []
    results = []
    for pdf in pdfs:
        try:
            results.append(process_pdf(str(pdf)))
        except Exception as e:
            print(f"  ✗ Failed {pdf.name}: {e}")
    return results


if __name__ == "__main__":
    target = sys.argv[1].strip() if len(sys.argv) > 1 else "Sample.pdf"
    try:
        if os.path.isdir(target):
            process_folder(target)
        else:
            process_pdf(target)
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        traceback.print_exc()
        sys.exit(1)